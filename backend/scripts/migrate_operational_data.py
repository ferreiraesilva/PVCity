import sys
from pathlib import Path

# Permite importar os módulos da app
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dotenv import load_dotenv

load_dotenv()

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db.base import Base
from app.db.session import engine
from app.models.enterprise import Enterprise
from app.models.real_estate_agency import RealEstateAgency
from app.models.unit_standard_flow import UnitStandardFlow
from app.services.label_normalizer import normalize_periodicity_label

WORKBOOK_SOURCE_DIR = (
    Path(__file__).resolve().parents[2] / "projects-docs" / "references" / "source-of-truth"
)
FLOW_SHEET = "Tabela Venda - Parcela"
IMOBS_SHEET = "Imobs"

SALE_ROW_SLOT_MAP = {
    "Sinal": 39,
    "Entrada": 40,
    "Mensais": 43,
    "Semestrais": 44,
    "Única": 45,
    "Unica": 45,
    "Anuais": 46,
    "Permuta": 54,
    "Veículo": 55,
    "Veiculo": 55,
    "Financ. Bancário": 56,
    "Financ. Bancario": 56,
    "Financ. Direto": 57,
}


def _to_float(value, default=0.0):
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int(value, default=0):
    if value is None:
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _row_slot_for(periodicity: str, fallback_slot: int) -> int:
    normalized = normalize_periodicity_label(periodicity) or periodicity
    return SALE_ROW_SLOT_MAP.get(normalized, fallback_slot)


def extract_sheet_rows(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]


def migrate():
    print("Sincronizando schema do banco...")
    Base.metadata.create_all(bind=engine)

    workbook_path = next(WORKBOOK_SOURCE_DIR.glob("*.xlsx"))
    print(f"Lendo Excel: {workbook_path.name}")
    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    flow_rows = extract_sheet_rows(wb, FLOW_SHEET)
    imobs_rows = extract_sheet_rows(wb, IMOBS_SHEET)
    wb.close()

    print(f"  Linhas de fluxo lidas: {len(flow_rows)}")
    print(f"  Imobiliárias lidas: {len(imobs_rows)}")

    with Session(engine) as session:
        existing_imobs = {item.name for item in session.query(RealEstateAgency.name).all()}
        new_imobs = []
        for row in imobs_rows:
            name = row.get("Imobiliaria") or row.get("Nome") or row.get("Imobiliária")
            if not name:
                continue
            normalized_name = str(name).strip()
            if normalized_name not in existing_imobs:
                new_imobs.append(RealEstateAgency(name=normalized_name, is_active=True))
                existing_imobs.add(normalized_name)

        if new_imobs:
            session.bulk_save_objects(new_imobs)
            print(f"  Imobiliárias inseridas: {len(new_imobs)}")

        enterprise_map = {enterprise.name: enterprise.id for enterprise in session.query(Enterprise).all()}

        new_flows = []
        skipped_flows = 0
        slots_counter = {}

        for row in flow_rows:
            enterprise_name = str(row.get("Nome da Origem") or "").strip()
            enterprise_id = enterprise_map.get(enterprise_name)
            if not enterprise_id:
                skipped_flows += 1
                continue

            periodicity = normalize_periodicity_label(row.get("Nomeclatura das Parcelas")) or ""
            if enterprise_id not in slots_counter:
                slots_counter[enterprise_id] = 39

            fallback_slot = slots_counter[enterprise_id]
            row_slot = _row_slot_for(periodicity, fallback_slot)

            new_flows.append(
                UnitStandardFlow(
                    enterprise_id=enterprise_id,
                    periodicity=periodicity,
                    installment_count=_to_int(row.get("N° Parcelas"), 1),
                    start_month=_to_int(row.get("Inicio Serie2"), 0),
                    installment_value=0.0,
                    percent=_to_float(row.get("Pcs")),
                    row_slot=row_slot,
                )
            )

            slots_counter[enterprise_id] = max(slots_counter[enterprise_id] + 1, row_slot + 1)

        if new_flows:
            affected_enterprises = list(slots_counter.keys())
            (
                session.query(UnitStandardFlow)
                .filter(UnitStandardFlow.enterprise_id.in_(affected_enterprises))
                .delete(synchronize_session=False)
            )
            session.bulk_save_objects(new_flows)
            print(f"  Linhas de fluxo inseridas: {len(new_flows)}")

        session.commit()
        print(f"  Fluxos ignorados por empreendimento ausente: {skipped_flows}")
        print("\nMigração operacional concluída.")


if __name__ == "__main__":
    migrate()
