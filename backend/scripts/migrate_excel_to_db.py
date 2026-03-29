"""
Script de migração única: Excel → Supabase  (versão otimizada com bulk insert)
Lê o arquivo Template PV e popula as tabelas `enterprise` e `unit`.

Uso:
    ..\.venv\Scripts\python.exe scripts/migrate_excel_to_db.py
"""
import sys
from pathlib import Path
from datetime import date, datetime

# Permite importar os módulos da app
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.db.session import engine
from app.models.enterprise import Enterprise
from app.models.unit import Unit

WORKBOOK_SOURCE_DIR = (
    Path(__file__).resolve().parents[2]
    / "projects-docs" / "references" / "source-of-truth"
)
PRODUCT_SHEET = "tbCadastroProduto"
REFERENCE_SHEET = "Referencias"


def _to_float(value, default=0.0):
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int(value):
    if value is None:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_iso(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def extract_sheet_rows(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]


def migrate():
    workbook_path = next(WORKBOOK_SOURCE_DIR.glob("*.xlsx"))
    print(f"Lendo (modo rápido): {workbook_path.name}")
    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    products = extract_sheet_rows(wb, PRODUCT_SHEET)
    references = extract_sheet_rows(wb, REFERENCE_SHEET)
    wb.close()

    print(f"  Empreendimentos lidos da planilha: {len(products)}")
    print(f"  Unidades lidas da planilha: {len(references)}")

    with Session(engine) as session:
        # --- Empreendimentos ---
        existing_enterprises = {e.name: e for e in session.query(Enterprise).all()}
        new_enterprises = []
        
        for row in products:
            name = row.get("Produto")
            if not name or str(name) in existing_enterprises:
                continue
            ent = Enterprise(
                name=str(name),
                work_code=str(row.get("Cod_obra") or ""),
                spe_name=str(row.get("SPE") or ""),
                vpl_rate_annual=_to_float(row.get("VPL")),
                discount_percent=_to_float(row.get("Descontos (%)")),
                delivery_month=_to_iso(row.get("Data de entrega")),
                launch_date=_to_iso(row.get("Data de Lancamento")),
                stage=str(row.get("Etapa") or ""),
                is_active=True,
            )
            new_enterprises.append(ent)

        if new_enterprises:
            session.bulk_save_objects(new_enterprises)
            session.flush()
            print(f"  ✅ Empreendimentos inseridos: {len(new_enterprises)}")
        else:
            print("  ⏭ Todos os empreendimentos já existem.")

        # Recarrega o mapa após bulk insert
        enterprise_map = {e.name: e for e in session.query(Enterprise).all()}

        # --- Unidades ---
        existing_keys = {
            u.product_unit_key
            for u in session.query(Unit.product_unit_key).all()
        }

        new_units = []
        skipped = 0
        no_enterprise = 0

        for row in references:
            enterprise_name = str(row.get("Nome da Origem") or "")
            unit_code = str(row.get("Unidade") or "")
            if not enterprise_name or not unit_code:
                continue

            product_unit_key = f"{enterprise_name}|{unit_code}"
            if product_unit_key in existing_keys:
                skipped += 1
                continue

            ent = enterprise_map.get(enterprise_name)
            if not ent:
                no_enterprise += 1
                continue

            new_units.append(Unit(
                enterprise_id=ent.id,
                code=unit_code,
                product_unit_key=product_unit_key,
                unit_type=str(row.get("Tipo") or ""),
                suites=_to_int(row.get("Suites")),
                garage_code=str(row.get("N° Escaninho") or ""),
                garage_spots=_to_int(row.get("Vagas de Garagem")),
                private_area_m2=_to_float(row.get("Area Privativa Total (m2)")),
                base_price=_to_float(row.get("Valor Total")),
                status=str(row.get("Status") or ""),
                ideal_capture_percent=1.0,
            ))

        if new_units:
            # Bulk insert em lotes de 500 para não estourar o pacote de rede
            batch_size = 500
            for i in range(0, len(new_units), batch_size):
                batch = new_units[i:i + batch_size]
                session.bulk_save_objects(batch)
                session.flush()
                print(f"  Lote {i // batch_size + 1}: {len(batch)} unidades inseridas...")

        session.commit()

        print(f"\n✅ Migração concluída!")
        print(f"   Empreendimentos no banco: {len(enterprise_map)}")
        print(f"   Unidades novas inseridas: {len(new_units)}")
        print(f"   Unidades já existentes (puladas): {skipped}")
        if no_enterprise:
            print(f"   ⚠ Unidades sem empreendimento correspondente: {no_enterprise}")


if __name__ == "__main__":
    migrate()
