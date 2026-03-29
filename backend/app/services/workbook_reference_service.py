from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from app.services.label_normalizer import (
    normalize_adjustment_label,
    normalize_modification_kind,
    normalize_periodicity_label,
    normalize_yes_no_label,
)


WORKBOOK_GLOB = "*.xlsx"
WORKBOOK_SOURCE_DIR = (
    Path(__file__).resolve().parents[3]
    / "projects-docs"
    / "references"
    / "source-of-truth"
)

ANALYSIS_SHEET = "Analise Proposta"
PRODUCT_SHEET = "tbCadastroProduto"
REFERENCE_SHEET = "Referencias"
PARCEL_SHEET = "Tabela Venda - Parcela"
AGENCY_SHEET = "Imobs"

PRODUCT_TABLE = "tbCadastroProduto"
REFERENCE_TABLE = "Tabela_Venda___Tabela"
PARCEL_TABLE = "Tabela_Venda___Parcela"
AGENCY_TABLE = "tbImobiliaria"

BASE_ROW_SLOT_MAP = {
    "Sinal": 20,
    "Entrada-1": 21,
    "Entrada-2": 22,
    "Entrada-3": 23,
    "Mensais": 24,
    "Semestrais": 25,
    "Unica": 26,
    "Financ. Bancario": 27,
}

SALE_ROW_SLOT_MAP = {
    "Sinal": 39,
    "Entrada-1": 40,
    "Entrada-2": 41,
    "Entrada-3": 42,
    "Mensais": 43,
    "Semestrais": 44,
    "Unica": 45,
    "Financ. Bancario": 56,
}


@dataclass(frozen=True)
class WorkbookTables:
    products: list[dict[str, Any]]
    references: list[dict[str, Any]]
    parcel_plan: list[dict[str, Any]]
    agencies: list[dict[str, Any]]
    defaults: dict[str, Any]


def _workbook_path() -> Path:
    try:
        return next(WORKBOOK_SOURCE_DIR.glob(WORKBOOK_GLOB))
    except StopIteration as exc:
        raise FileNotFoundError(
            f"No workbook found under {WORKBOOK_SOURCE_DIR}"
        ) from exc


def _to_iso(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _to_day_ten(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.replace(day=10).isoformat()
    return None


def _to_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_bool_sim_nao(value: Any) -> bool:
    return str(value).strip().lower() == "sim"


def _shift_month(start_value: str | None, offset: int) -> str | None:
    if not start_value:
        return None
    start_date = datetime.fromisoformat(start_value).date()
    year = start_date.year + ((start_date.month - 1 + offset) // 12)
    month = ((start_date.month - 1 + offset) % 12) + 1
    return date(year, month, 10).isoformat()


def _extract_table(ws: Any, table_name: str) -> list[dict[str, Any]]:
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows = list(
        ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
    )
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:]]


@lru_cache(maxsize=1)
def _load_workbook_tables() -> WorkbookTables:
    workbook = load_workbook(_workbook_path(), data_only=True)
    analysis_ws = workbook[ANALYSIS_SHEET]

    defaults = {
        "analysis_date": analysis_ws["E11"].value,
        "modification_kind": normalize_modification_kind(analysis_ws["D20"].value),
        "decorated_value_per_m2": analysis_ws["D21"].value,
        "facility_value_per_m2": analysis_ws["D22"].value,
        "prize_enabled": normalize_yes_no_label(analysis_ws["N4"].value),
        "fully_invoiced": normalize_yes_no_label(analysis_ws["N5"].value),
        "has_permuta": normalize_yes_no_label(analysis_ws["N6"].value),
        "primary_commission_label": analysis_ws["M8"].value,
        "primary_commission_percent": analysis_ws["N8"].value,
        "prize_commission_label": analysis_ws["M9"].value,
        "prize_commission_percent": analysis_ws["N9"].value,
        "city_sales_manager_name": analysis_ws["K9"].value,
        "real_estate_name": analysis_ws["K10"].value,
        "broker_name": analysis_ws["K11"].value,
        "manager_name": analysis_ws["K12"].value,
    }

    return WorkbookTables(
        products=_extract_table(workbook[PRODUCT_SHEET], PRODUCT_TABLE),
        references=_extract_table(workbook[REFERENCE_SHEET], REFERENCE_TABLE),
        parcel_plan=_extract_table(workbook[PARCEL_SHEET], PARCEL_TABLE),
        agencies=_extract_table(workbook[AGENCY_SHEET], AGENCY_TABLE),
        defaults=defaults,
    )


def _find_product(enterprise_name: str) -> dict[str, Any]:
    tables = _load_workbook_tables()
    for row in tables.products:
        if str(row.get("Produto")) == enterprise_name:
            return row
    raise KeyError(f"Unknown enterprise: {enterprise_name}")


def _find_reference(product_unit_key: str) -> dict[str, Any]:
    tables = _load_workbook_tables()
    for row in tables.references:
        if str(row.get("Id")) == product_unit_key:
            return row
    raise KeyError(f"Unknown product unit key: {product_unit_key}")


def _find_parcel_row(enterprise_name: str, parcel_name: str) -> dict[str, Any]:
    tables = _load_workbook_tables()
    for row in tables.parcel_plan:
        if (
            str(row.get("Nome da Origem")) == enterprise_name
            and normalize_periodicity_label(row.get("Nomeclatura das Parcelas"))
            == normalize_periodicity_label(parcel_name)
        ):
            return row
    raise KeyError(f"Unknown parcel plan row: {enterprise_name}|{parcel_name}")


def _commission_defaults() -> dict[str, float | str | None]:
    defaults = _load_workbook_tables().defaults
    return {
        "primary_commission_label": defaults["primary_commission_label"],
        "primary_commission_percent": _to_float(defaults["primary_commission_percent"]),
        "prize_commission_label": defaults["prize_commission_label"],
        "prize_commission_percent": _to_float(defaults["prize_commission_percent"]),
    }


def _total_commission_percent(prize_enabled: bool, has_permuta: bool) -> float:
    commission_defaults = _commission_defaults()
    base_percent = _to_float(commission_defaults["primary_commission_percent"])
    prize_percent = (
        _to_float(commission_defaults["prize_commission_percent"]) if prize_enabled else 0.0
    )
    # Workbook adds a proportional delta when permuta is active, but the exact
    # value depends on the exchange grid. Defaults stay on the baseline rate.
    if has_permuta:
        return base_percent + prize_percent
    return base_percent + prize_percent


def _product_context_payload(
    enterprise_name: str,
    unit_code: str,
    reference_row: dict[str, Any],
    product_row: dict[str, Any],
) -> dict[str, Any]:
    defaults = _load_workbook_tables().defaults
    analysis_date = defaults["analysis_date"]
    private_area = _to_float(reference_row.get("Area Privativa Total (m2)"))

    return {
        "enterprise_name": enterprise_name,
        "unit_code": str(unit_code),
        "product_unit_key": f"{enterprise_name}|{unit_code}",
        "garage_code": reference_row.get("N° Escaninho"),
        "private_area_m2": private_area,
        "delivery_month": _to_iso(product_row.get("Data de entrega")),
        "default_analysis_date_kind": "server_today",
        "default_modification_kind": defaults["modification_kind"],
        "default_decorated_value_per_m2": _to_float(defaults["decorated_value_per_m2"]),
        "default_facility_value_per_m2": _to_float(defaults["facility_value_per_m2"]),
        "default_area_for_modification_m2": private_area,
        "default_prize_enabled": _to_bool_sim_nao(defaults["prize_enabled"]),
        "default_fully_invoiced": _to_bool_sim_nao(defaults["fully_invoiced"]),
        "default_has_permuta": False,
        "default_analysis_date": _to_iso(analysis_date),
        "base_price": _to_float(reference_row.get("Valor Total")),
        "enterprise_discount_percent": _to_float(product_row.get("Descontos (%)")),
        "enterprise_vpl_rate_annual": _to_float(product_row.get("VPL")),
    }


def _base_row_payload(
    row_slot: int,
    periodicity: str,
    installment_count: float,
    start_month: str | None,
    installment_value: float,
    percent: float,
    total_vgv: float,
    commission_target_value: float,
    commission_paid_value: float,
) -> dict[str, Any]:
    return {
        "row_slot": row_slot,
        "installment_count": installment_count,
        "periodicity": normalize_periodicity_label(periodicity),
        "start_month": start_month,
        "installment_value": installment_value,
        "percent": percent,
        "total_vgv": total_vgv,
        "commission_target_value": commission_target_value,
        "commission_paid_value": commission_paid_value,
        "net_value": installment_value - commission_paid_value,
    }


def _sale_row_payload(
    row_slot: int,
    periodicity: str,
    installment_count: float,
    start_month: str | None,
    installment_value: float,
    percent: float,
    total_vgv: float,
    adjustment_type: str,
) -> dict[str, Any]:
    return {
        "row_slot": row_slot,
        "installment_count": installment_count,
        "periodicity": normalize_periodicity_label(periodicity),
        "start_month": start_month,
        "installment_value": installment_value,
        "percent": percent,
        "total_vgv": total_vgv,
        "adjustment_type": normalize_adjustment_label(adjustment_type),
        "notes": None,
    }


def _build_base_and_sale_rows(
    enterprise_name: str,
    total_vgv: float,
    prize_enabled: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    total_commission_value = total_vgv * _total_commission_percent(
        prize_enabled=prize_enabled,
        has_permuta=False,
    )

    signal_plan = _find_parcel_row(enterprise_name, "Sinal")
    entry_plan = _find_parcel_row(enterprise_name, "Entrada")
    monthly_plan = _find_parcel_row(enterprise_name, "Mensais")
    semiannual_plan = _find_parcel_row(enterprise_name, "Semestrais")
    unique_plan = _find_parcel_row(enterprise_name, "Única")
    financing_plan = _find_parcel_row(enterprise_name, "Financ. Bancário")

    base_rows: list[dict[str, Any]] = []
    sale_rows: list[dict[str, Any]] = []

    def plan_value(plan: dict[str, Any]) -> tuple[float, float, float, str | None]:
        count = max(_to_float(plan.get("N° Parcelas")), 0.0)
        percent = _to_float(plan.get("Pcs"))
        installment_value = total_vgv * percent
        return count, percent, installment_value, _to_day_ten(plan.get("Mês de Inicio"))

    signal_count, signal_percent, signal_value, signal_start = plan_value(signal_plan)
    signal_total = signal_count * signal_value
    signal_target = total_commission_value * 0.7
    signal_paid = min(signal_total - total_vgv * 0.03, signal_target)
    commission_carry = signal_target - signal_paid
    base_rows.append(
        _base_row_payload(
            row_slot=BASE_ROW_SLOT_MAP["Sinal"],
            periodicity="Sinal",
            installment_count=signal_count,
            start_month=signal_start,
            installment_value=signal_value,
            percent=signal_percent,
            total_vgv=signal_total,
            commission_target_value=signal_target,
            commission_paid_value=signal_paid,
        )
    )
    sale_rows.append(
        _sale_row_payload(
            row_slot=SALE_ROW_SLOT_MAP["Sinal"],
            periodicity="Sinal",
            installment_count=signal_count,
            start_month=signal_start,
            installment_value=signal_value,
            percent=signal_percent,
            total_vgv=signal_total,
            adjustment_type="Fixas Irreajustaveis",
        )
    )

    entry_count, entry_percent, entry_value, entry_start = plan_value(entry_plan)
    entry_target = total_commission_value * 0.1
    for index in range(3):
        row_key = f"Entrada-{index + 1}"
        paid = min(entry_value, entry_target + commission_carry)
        commission_carry = commission_carry + entry_target - paid
        start_value = _shift_month(entry_start, index)
        base_rows.append(
            _base_row_payload(
                row_slot=BASE_ROW_SLOT_MAP[row_key],
                periodicity="Entrada",
                installment_count=1.0 if entry_count else 0.0,
                start_month=start_value,
                installment_value=entry_value,
                percent=entry_percent,
                total_vgv=entry_value,
                commission_target_value=entry_target,
                commission_paid_value=paid,
            )
        )
        sale_rows.append(
            _sale_row_payload(
                row_slot=SALE_ROW_SLOT_MAP[row_key],
                periodicity="Entrada",
                installment_count=1.0 if entry_count else 0.0,
                start_month=start_value,
                installment_value=entry_value,
                percent=entry_percent,
                total_vgv=entry_value,
                adjustment_type="Fixas Irreajustaveis",
            )
        )

    for parcel_name, normalized_name, adjustment_type in [
        ("Mensais", "Mensais", "INCC"),
        ("Semestrais", "Semestrais", "INCC"),
        ("Única", "Unica", "INCC"),
        ("Financ. Bancário", "Financ. Bancario", "IGPM + 12% a.a"),
    ]:
        plan = {
            "Mensais": monthly_plan,
            "Semestrais": semiannual_plan,
            "Única": unique_plan,
            "Financ. Bancário": financing_plan,
        }[parcel_name]
        count, percent, installment_value, start_month = plan_value(plan)
        total_row_vgv = installment_value * count
        base_rows.append(
            _base_row_payload(
                row_slot=BASE_ROW_SLOT_MAP[normalized_name],
                periodicity=parcel_name,
                installment_count=count,
                start_month=start_month,
                installment_value=installment_value,
                percent=percent,
                total_vgv=total_row_vgv,
                commission_target_value=0.0,
                commission_paid_value=0.0,
            )
        )
        sale_rows.append(
            _sale_row_payload(
                row_slot=SALE_ROW_SLOT_MAP[normalized_name],
                periodicity=parcel_name,
                installment_count=count,
                start_month=start_month,
                installment_value=installment_value,
                percent=percent,
                total_vgv=total_row_vgv,
                adjustment_type=adjustment_type,
            )
        )

    return base_rows, sale_rows


def build_permuta_rows_from_sale_rows(
    sale_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    exchange_rows: list[dict[str, Any]] = []
    for row in sale_rows:
        transformed = dict(row)
        if normalize_periodicity_label(row.get("periodicity")) == "Semestrais":
            installment_value = _to_float(row.get("installment_value"))
            transformed["periodicity"] = "Permuta"
            transformed["installment_count"] = 1.0
            transformed["total_vgv"] = installment_value
            transformed["notes"] = "Derived from Semestrais for the permuta scenario"
        exchange_rows.append(transformed)
    return exchange_rows


class WorkbookReferenceService:
    """
    Reverse-engineering and parity helper.
    Do not use as an operational dependency for the application runtime.
    """

    def get_reference_data(self) -> dict[str, Any]:
        tables = _load_workbook_tables()
        products = [
            {
                "enterprise_name": row.get("Produto"),
                "work_code": row.get("Cod_obra"),
                "spe_name": row.get("SPE"),
                "default_discount_percent": _to_float(row.get("Descontos (%)")),
                "default_vpl_rate_annual": _to_float(row.get("VPL")),
                "delivery_month": _to_iso(row.get("Data de entrega")),
                "launch_date": _to_iso(row.get("Data de Lancamento")),
                "stage": row.get("Etapa"),
                "personalization_status": row.get("Personalização"),
                "personalization_deadline": _to_iso(row.get("Data Personalização Encerramento")),
            }
            for row in tables.products
        ]
        products.sort(key=lambda row: (str(row["enterprise_name"] or ""), str(row["work_code"] or "")))

        unit_lookup_keys = [
            {
                "product_unit_key": row.get("Id"),
                "enterprise_name": row.get("Nome da Origem"),
                "unit_code": str(row.get("Unidade")),
                "unit_type": row.get("Tipo"),
                "suites": row.get("Suites"),
                "private_area_m2": _to_float(row.get("Area Privativa Total (m2)")),
                "garage_spots": row.get("Vagas de Garagem"),
                "base_price": _to_float(row.get("Valor Total")),
                "status": row.get("Status"),
            }
            for row in tables.references
        ]
        unit_lookup_keys.sort(
            key=lambda row: (
                str(row["enterprise_name"] or ""),
                str(row["unit_code"] or ""),
            )
        )


        agencies = [
            {
                "name": row.get("Imobiliaria"),
                "manager_cv_name": row.get("GerenteCV"),
            }
            for row in tables.agencies
        ]

        return {
            "products": products,
            "unit_lookup_keys": unit_lookup_keys,
            "real_estate_agencies": agencies,
            "enums": {
                "boolean_ptbr": ["Sim", "Não"],
                "modification_kind": ["Não", "Decorado (R$/m²)", "Facility (R$/m²)"],
                "periodicity": [
                    "Sinal",
                    "Entrada",
                    "Mensais",
                    "Semestrais",
                    "Única",
                    "Permuta",
                    "Anuais",
                    "Veículo",
                    "Financ. Bancário",
                    "Financ. Direto",
                ],
                "financing_kind": ["Financ. Bancário", "Financ. Direto"],
                "adjustment_type": [
                    "Fixas Irreajustaveis",
                    "INCC",
                    "IGPM + 12% a.a",
                    "IPCA + 0,99% a.m",
                    "IPCA + 13,65% a.a",
                ],
            },
        }

    def get_unit_defaults(self, enterprise_name: str, unit_code: str) -> dict[str, Any]:
        reference_row = _find_reference(f"{enterprise_name}|{unit_code}")
        product_row = _find_product(enterprise_name)
        prize_enabled = _to_bool_sim_nao(_load_workbook_tables().defaults["prize_enabled"])

        base_rows, sale_rows = _build_base_and_sale_rows(
            enterprise_name=enterprise_name,
            total_vgv=_to_float(reference_row.get("Valor Total")),
            prize_enabled=prize_enabled,
        )

        return {
            "product_context": _product_context_payload(
                enterprise_name=enterprise_name,
                unit_code=unit_code,
                reference_row=reference_row,
                product_row=product_row,
            ),
            "base_sale_table_rows": base_rows,
            "default_sale_flow_rows": sale_rows,
            "default_exchange_flow_rows": build_permuta_rows_from_sale_rows(sale_rows),
            "commercial_context": {
                "city_sales_manager_name": _load_workbook_tables().defaults["city_sales_manager_name"],
                "real_estate_name": _load_workbook_tables().defaults["real_estate_name"],
                "broker_name": _load_workbook_tables().defaults["broker_name"],
                "manager_name": _load_workbook_tables().defaults["manager_name"],
            },
            "commission_context": {
                **_commission_defaults(),
                "secondary_commission_slots": [
                    {"slot": 1, "label": None, "percent": None},
                    {"slot": 2, "label": None, "percent": None},
                ],
            },
        }

    def get_product_reference(self, enterprise_name: str, unit_code: str) -> dict[str, Any]:
        reference_row = _find_reference(f"{enterprise_name}|{unit_code}")
        product_row = _find_product(enterprise_name)
        return {
            "reference_row": reference_row,
            "product_row": product_row,
        }
